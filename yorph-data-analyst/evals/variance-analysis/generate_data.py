"""
Generate synthetic dirty FP&A data for the variance analysis eval.
Run once to create the data files, then delete this script if desired.
"""

import csv
import random
import datetime
import os

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl

random.seed(42)

OUT_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(OUT_DIR, exist_ok=True)

# --- Shared reference data ---

CATEGORIES = ["Electronics", "Apparel", "Home & Kitchen", "Sports", "Beauty", "Office Supplies"]
# Dirty versions that will appear in different systems
CATEGORY_DIRTY = {
    "Electronics": ["Electronics", "electronics", "ELECTRONICS"],
    "Apparel": ["Apparel", "apparel", "APPAREL"],
    "Home & Kitchen": ["Home & Kitchen", "home & kitchen", "Home and Kitchen"],
    "Sports": ["Sports", "sports", "SPORTS"],
    "Beauty": ["Beauty", "beauty", "BEAUTY"],
    "Office Supplies": ["Office Supplies", "office supplies", "Office supplies"],
}

SUBCATEGORIES = {
    "Electronics": ["Laptops", "Phones", "Tablets", "Accessories", "Audio"],
    "Apparel": ["Men's", "Women's", "Kids", "Footwear", "Outerwear"],
    "Home & Kitchen": ["Cookware", "Furniture", "Decor", "Appliances", "Storage"],
    "Sports": ["Fitness", "Outdoor", "Team Sports", "Water Sports", "Cycling"],
    "Beauty": ["Skincare", "Makeup", "Haircare", "Fragrance", "Tools"],
    "Office Supplies": ["Paper", "Writing", "Desk Organization", "Tech Accessories", "Breakroom"],
}

REGIONS = ["North America", "Europe", "Asia Pacific", "Latin America"]
CURRENCIES_BY_REGION = {
    "North America": "USD",
    "Europe": "EUR",
    "Asia Pacific": "JPY",
    "Latin America": "BRL",
}
# Dirty currency representations for USD
USD_VARIANTS = ["USD", "usd", "US Dollar", "USD", "USD"]  # weighted toward USD

CHANNEL_TYPES = ["Online", "Retail", "Wholesale", "Marketplace"]

RETURN_FLAG_TRUE = ["Y", "yes", "1", "TRUE", "Y"]
RETURN_FLAG_FALSE = ["N", "no", "0", "FALSE", "N"]

# Generate 500 products
products = []
for i in range(1, 501):
    cat = random.choice(CATEGORIES)
    pid = f"PRD-{i:04d}"
    products.append({
        "product_id": pid,
        "product_name": f"Product {pid}",
        "category": cat,
        "subcategory": random.choice(SUBCATEGORIES[cat]),
        "launch_date": (datetime.date(2020, 1, 1) + datetime.timedelta(days=random.randint(0, 1400))).isoformat(),
        "status": random.choice(["Active"] * 9 + ["Discontinued"]),
    })

product_ids = [p["product_id"] for p in products]
# 20 extra "ghost" product IDs that exist in transactions but NOT in catalog
ghost_product_ids = [f"PRD-{i:04d}" for i in range(501, 521)]

# Generate 50 channels
channels = []
channel_id_counter = 1
for ct in CHANNEL_TYPES:
    for j in range(1, 13):
        cid = f"CH-{channel_id_counter:03d}"
        parent = None
        if j > 3:
            parent = f"CH-{channel_id_counter - 3:03d}"
        channels.append({
            "channel_id": cid,
            "channel_name": f"{ct} Channel {j}",
            "channel_type": ct,
            "parent_channel": parent or "",
            "commission_rate": round(random.uniform(0.02, 0.15), 4),
        })
        channel_id_counter += 1
# Pad to ~50
while len(channels) < 50:
    cid = f"CH-{channel_id_counter:03d}"
    ct = random.choice(CHANNEL_TYPES)
    channels.append({
        "channel_id": cid,
        "channel_name": f"{ct} Extra {channel_id_counter}",
        "channel_type": ct,
        "parent_channel": "",
        "commission_rate": round(random.uniform(0.02, 0.15), 4),
    })
    channel_id_counter += 1

valid_channel_ids = [c["channel_id"] for c in channels]
# Ghost channel IDs that will appear in transactions but not in channels.csv
ghost_channel_ids = [f"CH-{i:03d}" for i in range(channel_id_counter, channel_id_counter + 8)]

# --- 1. channels.csv (clean, only orphan refs as dirty trick) ---

with open(os.path.join(OUT_DIR, "channels.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["channel_id", "channel_name", "channel_type", "parent_channel", "commission_rate"])
    w.writeheader()
    w.writerows(channels)

print(f"channels.csv: {len(channels)} rows")

# --- 2. fx_rates.csv (clean, no dirty tricks) ---

fx_pairs = [("EUR", "USD"), ("JPY", "USD"), ("BRL", "USD"), ("GBP", "USD")]
base_rates = {"EUR": 1.08, "JPY": 0.0067, "BRL": 0.20, "GBP": 1.27}

fx_rows = []
start = datetime.date(2024, 1, 1)
end = datetime.date(2024, 12, 31)
d = start
while d <= end:
    for from_c, to_c in fx_pairs:
        rate = base_rates[from_c] * (1 + random.gauss(0, 0.005))
        fx_rows.append({
            "date": d.isoformat(),
            "from_currency": from_c,
            "to_currency": to_c,
            "rate": round(rate, 6),
        })
    d += datetime.timedelta(days=1)

with open(os.path.join(OUT_DIR, "fx_rates.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["date", "from_currency", "to_currency", "rate"])
    w.writeheader()
    w.writerows(fx_rows)

print(f"fx_rates.csv: {len(fx_rows)} rows")

# --- 3. transactions.csv (~30K rows, lots of dirty tricks) ---

def dirty_product_id(pid):
    """5% chance of trailing whitespace or case mismatch"""
    if random.random() < 0.05:
        choice = random.choice(["trailing_space", "lowercase", "uppercase"])
        if choice == "trailing_space":
            return pid + "  "
        elif choice == "lowercase":
            return pid.lower()
        else:
            return pid.upper()
    return pid

def dirty_date(d):
    """Mix date formats"""
    if random.random() < 0.3:
        return d.strftime("%m/%d/%Y")
    return d.isoformat()

def dirty_discount():
    """Mix decimal and whole-number discount formats"""
    base = random.choice([0, 0, 0, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30])
    if base == 0:
        return 0
    if random.random() < 0.4:
        return int(base * 100)  # whole number like 15
    return base  # decimal like 0.15

def dirty_return_flag(is_return):
    if random.random() < 0.03:
        return ""  # null
    if is_return:
        return random.choice(RETURN_FLAG_TRUE)
    return random.choice(RETURN_FLAG_FALSE)

def dirty_currency(region):
    currency = CURRENCIES_BY_REGION[region]
    if currency == "USD":
        return random.choice(USD_VARIANTS)
    return currency

txn_rows = []
txn_id_counter = 1
duplicate_txn_ids = set()

for _ in range(30000):
    txn_id = f"TXN-{txn_id_counter:06d}"
    txn_id_counter += 1

    d = datetime.date(2024, 1, 1) + datetime.timedelta(days=random.randint(0, 365))
    region = random.choice(REGIONS)

    # 0.5% chance to use a ghost product
    if random.random() < 0.005:
        pid = random.choice(ghost_product_ids)
    else:
        pid = random.choice(product_ids)

    # 2% chance to use a ghost channel
    if random.random() < 0.02:
        cid = random.choice(ghost_channel_ids)
    else:
        cid = random.choice(valid_channel_ids)

    qty = random.randint(1, 50)
    # 0.3% chance of negative qty (data entry error, not a return)
    if random.random() < 0.003:
        qty = -random.randint(1, 10)

    is_return = random.random() < 0.08
    unit_price = round(random.uniform(5, 500), 2)

    row = {
        "txn_id": txn_id,
        "date": dirty_date(d),
        "product_id": dirty_product_id(pid),
        "channel_id": cid,
        "qty": qty,
        "unit_price": unit_price,
        "currency": dirty_currency(region),
        "discount_pct": dirty_discount(),
        "return_flag": dirty_return_flag(is_return),
        "region": region,
    }
    txn_rows.append(row)

    # ~200 duplicates with slightly different amounts
    if len(duplicate_txn_ids) < 200 and random.random() < 0.007:
        duplicate_txn_ids.add(txn_id)
        dup_row = dict(row)
        dup_row["unit_price"] = round(row["unit_price"] * random.uniform(0.98, 1.02), 2)
        dup_row["qty"] = row["qty"] + random.choice([-1, 0, 1])
        txn_rows.append(dup_row)

random.shuffle(txn_rows)

with open(os.path.join(OUT_DIR, "transactions.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["txn_id", "date", "product_id", "channel_id", "qty", "unit_price", "currency", "discount_pct", "return_flag", "region"])
    w.writeheader()
    w.writerows(txn_rows)

print(f"transactions.csv: {len(txn_rows)} rows ({len(duplicate_txn_ids)} duplicate txn_ids)")

# --- 4. products.xlsx (2 sheets) ---

wb = openpyxl.Workbook()

# Sheet 1: Catalog
ws_catalog = wb.active
ws_catalog.title = "Catalog"
ws_catalog.append(["product_id", "product_name", "category", "subcategory", "launch_date", "status"])
for p in products:
    # Dirty: inconsistent category capitalization
    dirty_cat = random.choice(CATEGORY_DIRTY[p["category"]])
    ws_catalog.append([p["product_id"], p["product_name"], dirty_cat, p["subcategory"], p["launch_date"], p["status"]])

# Sheet 2: Price History (~2K rows)
ws_prices = wb.create_sheet("Price History")
ws_prices.append(["product_id", "effective_date", "list_price", "cost"])

price_rows = 0
for p in products:
    # 2-6 price changes per product
    n_changes = random.randint(2, 6)
    dates = sorted([datetime.date(2023, 1, 1) + datetime.timedelta(days=random.randint(0, 730)) for _ in range(n_changes)])

    for i, ed in enumerate(dates):
        list_price = round(random.uniform(10, 600), 2)
        cost = round(list_price * random.uniform(0.3, 0.7), 2)

        # 2% chance of bad cost (0 or negative)
        if random.random() < 0.02:
            cost = random.choice([0, -round(random.uniform(1, 50), 2)])

        ws_prices.append([p["product_id"], ed.isoformat(), list_price, cost])
        price_rows += 1

        # 3% chance of overlapping date range (add another entry within 5 days)
        if random.random() < 0.03 and i < n_changes - 1:
            overlap_date = ed + datetime.timedelta(days=random.randint(0, 5))
            ws_prices.append([p["product_id"], overlap_date.isoformat(), round(list_price * 1.05, 2), cost])
            price_rows += 1

wb.save(os.path.join(OUT_DIR, "products.xlsx"))
print(f"products.xlsx: Catalog={len(products)} rows, Price History={price_rows} rows")

# --- 5. budget_plan.xlsx (2 sheets) ---

wb2 = openpyxl.Workbook()

# Sheet 1: Annual budget at category x channel_type x region grain
ws_annual = wb2.active
ws_annual.title = "Annual"
ws_annual.append(["category", "channel_type", "region", "budget_revenue", "budget_units", "budget_cogs"])

budget_rows = 0
# Use slightly different category names than the product catalog (dirty)
BUDGET_CATEGORIES = {
    "Electronics": "Consumer Electronics",
    "Apparel": "Apparel & Clothing",
    "Home & Kitchen": "Home & Kitchen",  # this one matches
    "Sports": "Sporting Goods",
    "Beauty": "Beauty & Personal Care",
    "Office Supplies": "Office",
}

for cat in CATEGORIES:
    budget_cat = BUDGET_CATEGORIES[cat]
    for ct in CHANNEL_TYPES:
        for region in REGIONS:
            rev = round(random.uniform(50000, 2000000), 2)
            units = random.randint(500, 50000)
            cogs = round(rev * random.uniform(0.4, 0.7), 2)
            ws_annual.append([budget_cat, ct, region, rev, units, cogs])
            budget_rows += 1

# Sheet 2: Quarterly phasing
ws_quarterly = wb2.create_sheet("Quarterly")
ws_quarterly.append(["category", "quarter", "phasing_pct"])

for cat in CATEGORIES:
    budget_cat = BUDGET_CATEGORIES[cat]
    # Generate 4 quarters, but some won't sum to 100%
    phasings = [random.uniform(15, 35) for _ in range(4)]

    # 30% chance phasing doesn't sum to 100
    if random.random() < 0.3:
        total = sum(phasings)
        phasings = [round(p / total * 100, 1) for p in phasings]
        # Nudge one value to break the sum
        phasings[random.randint(0, 3)] += round(random.uniform(-3, 3), 1)
    else:
        total = sum(phasings)
        phasings = [round(p / total * 100, 1) for p in phasings]
        # Fix rounding so it sums to 100
        phasings[3] = round(100 - sum(phasings[:3]), 1)

    for q in range(1, 5):
        ws_quarterly.append([budget_cat, f"Q{q}", phasings[q - 1]])

wb2.save(os.path.join(OUT_DIR, "budget_plan.xlsx"))
print(f"budget_plan.xlsx: Annual={budget_rows} rows, Quarterly={len(CATEGORIES) * 4} rows")

print("\nAll files generated successfully!")
